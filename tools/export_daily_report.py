#!/usr/bin/env python3
"""
Export Firestore order data to a multi-sheet .xlsx report.

Read-only. Uses gcloud token (no service account file).

Usage:
  python3 tools/export_daily_report.py                       # last 7 days (Panamá TZ)
  python3 tools/export_daily_report.py --date 2026-04-13     # single day
  python3 tools/export_daily_report.py --from 2026-04-01 --to 2026-04-16
  python3 tools/export_daily_report.py --out my-report.xlsx  # custom path

Sheets:
  Summary       : totals + payment method breakdown
  Daily         : one row per day in range
  Orders        : one row per order
  Items         : one row per order_item
  TopProducts   : top 25 products by revenue (excluding cancelled items)
"""

# Bootstrap: load openpyxl from repo .venv if present
import os
import sys
_REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_VENV_LIB = os.path.join(_REPO, ".venv", "lib")
if os.path.isdir(_VENV_LIB):
    for d in os.listdir(_VENV_LIB):
        sp = os.path.join(_VENV_LIB, d, "site-packages")
        if os.path.isdir(sp) and sp not in sys.path:
            sys.path.insert(0, sp)

import argparse
import json
import subprocess
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import datetime, timedelta, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("openpyxl not found. Install: ./.venv/bin/pip install openpyxl")

PROJECT_ID = "restaurant-os-68c79"
ORG_ID     = "demo-org"
BRANCH_ID  = "demo-branch"
PANAMA_TZ  = timezone(timedelta(hours=-5))   # America/Panama, no DST
FIRESTORE_BASE = f"https://firestore.googleapis.com/v1/projects/{PROJECT_ID}/databases/(default)/documents"


# ---------- HTTP / Firestore helpers ----------

def get_token():
    r = subprocess.run(["gcloud", "auth", "print-access-token"],
                       capture_output=True, text=True)
    return r.stdout.strip()


def http(method, url, token, body=None):
    data = None
    headers = {"Authorization": f"Bearer {token}"}
    if body is not None:
        data = json.dumps(body).encode()
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode()
            return resp.status, json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        return e.code, {"error": e.read().decode()[:1000]}


def parse_value(v):
    if "stringValue" in v:    return v["stringValue"]
    if "integerValue" in v:   return int(v["integerValue"])
    if "doubleValue" in v:    return v["doubleValue"]
    if "booleanValue" in v:   return v["booleanValue"]
    if "nullValue" in v:      return None
    if "timestampValue" in v: return v["timestampValue"]
    if "arrayValue" in v:
        return [parse_value(x) for x in v["arrayValue"].get("values", [])]
    if "mapValue" in v:
        return {k: parse_value(val) for k, val in v["mapValue"].get("fields", {}).items()}
    return None


def parse_doc(doc):
    name = doc.get("name", "")
    out = {"_id": name.split("/")[-1] if name else "?"}
    for k, v in doc.get("fields", {}).items():
        out[k] = parse_value(v)
    return out


def list_all(collection, token):
    docs, page_token = [], None
    while True:
        url = f"{FIRESTORE_BASE}/{collection}?pageSize=300"
        if page_token:
            url += f"&pageToken={page_token}"
        st, data = http("GET", url, token)
        if st != 200:
            sys.exit(f"ERROR listing {collection}: {data}")
        docs.extend(data.get("documents", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    return docs


def run_query(structured_query, token):
    url = f"{FIRESTORE_BASE}:runQuery"
    st, data = http("POST", url, token, {"structuredQuery": structured_query})
    if st != 200:
        sys.exit(f"ERROR runQuery: {data}")
    return [parse_doc(item["document"]) for item in data if "document" in item]


# ---------- Date / TZ ----------

def parse_date(s):
    return datetime.strptime(s, "%Y-%m-%d").date()


def day_bounds_utc(d):
    """Return (start_utc_iso, end_utc_iso) for full Panamá day d."""
    start = datetime(d.year, d.month, d.day, 0, 0, 0, tzinfo=PANAMA_TZ).astimezone(timezone.utc)
    end   = datetime(d.year, d.month, d.day, 23, 59, 59, 999000, tzinfo=PANAMA_TZ).astimezone(timezone.utc)
    fmt = "%Y-%m-%dT%H:%M:%S.%fZ"
    return start.strftime(fmt), end.strftime(fmt)


def range_bounds_utc(start_d, end_d):
    s, _ = day_bounds_utc(start_d)
    _, e = day_bounds_utc(end_d)
    return s, e


def to_panama_dt(iso_ts):
    """Parse Firestore timestamp string into a Panamá-local datetime."""
    if not iso_ts:
        return None
    s = iso_ts.replace("Z", "+00:00")
    # Firestore can return up to 9 fractional digits (nanos); fromisoformat caps at 6
    if "." in s:
        head, frac = s.split(".", 1)
        if "+" in frac:
            f, off = frac.split("+", 1); s = f"{head}.{f[:6]}+{off}"
        elif "-" in frac:
            f, off = frac.split("-", 1); s = f"{head}.{f[:6]}-{off}"
        else:
            s = f"{head}.{frac[:6]}"
    return datetime.fromisoformat(s).astimezone(PANAMA_TZ)


# ---------- Firestore queries ----------

def fetch_orders(start_d, end_d, token):
    s, e = range_bounds_utc(start_d, end_d)
    sq = {
        "from": [{"collectionId": "orders"}],
        "where": {
            "compositeFilter": {
                "op": "AND",
                "filters": [
                    {"fieldFilter": {"field": {"fieldPath": "orgId"},
                                     "op": "EQUAL",
                                     "value": {"stringValue": ORG_ID}}},
                    {"fieldFilter": {"field": {"fieldPath": "createdAt"},
                                     "op": "GREATER_THAN_OR_EQUAL",
                                     "value": {"timestampValue": s}}},
                    {"fieldFilter": {"field": {"fieldPath": "createdAt"},
                                     "op": "LESS_THAN_OR_EQUAL",
                                     "value": {"timestampValue": e}}},
                ]
            }
        },
        "orderBy": [{"field": {"fieldPath": "createdAt"},
                     "direction": "DESCENDING"}],
    }
    orders = run_query(sq, token)
    orders.sort(key=lambda o: o.get("createdAt") or "")
    return orders


def fetch_items_for_orders(order_ids, token):
    items = []
    BATCH = 30  # Firestore IN limit
    for i in range(0, len(order_ids), BATCH):
        chunk = order_ids[i:i+BATCH]
        sq = {
            "from": [{"collectionId": "order_items"}],
            "where": {
                "fieldFilter": {
                    "field": {"fieldPath": "orderId"},
                    "op": "IN",
                    "value": {"arrayValue": {
                        "values": [{"stringValue": o} for o in chunk]
                    }}
                }
            }
        }
        items.extend(run_query(sq, token))
    return items


# ---------- Excel ----------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
LABEL_FONT  = Font(bold=True)
MONEY_FMT   = '"$"#,##0.00'

def write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def build_workbook(orders, items, cat_map, start_d, end_d, out_path):
    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    closed    = [o for o in orders if o.get("status") == "closed"]
    cancelled = [o for o in orders if o.get("status") == "cancelled"]
    subtotal  = sum(o.get("subtotal",  0) or 0 for o in orders)
    tax       = sum(o.get("taxAmount", 0) or 0 for o in orders)
    tip       = sum(o.get("tipAmount", 0) or 0 for o in orders)
    total     = sum(o.get("total",     0) or 0 for o in orders)
    avg_ticket = (total / len(closed)) if closed else 0.0

    by_method = defaultdict(float)
    for o in orders:
        pay = o.get("payment") or {}
        m = pay.get("method") or "(sin pago)"
        by_method[m] += o.get("total", 0) or 0

    rows = [
        ("Rango",                 f"{start_d} → {end_d} (zona Panamá)"),
        ("Pedidos totales",       len(orders)),
        ("Pedidos cerrados",      len(closed)),
        ("Pedidos cancelados",    len(cancelled)),
        ("Subtotal (sin tax/tip)", float(subtotal)),
        ("Impuestos",             float(tax)),
        ("Propinas",              float(tip)),
        ("Total facturado",       float(total)),
        ("Ticket promedio",       float(avg_ticket)),
        ("", ""),
        ("Pago: efectivo",        float(by_method.get("cash", 0))),
        ("Pago: tarjeta",         float(by_method.get("card", 0))),
        ("Pago: yappy",           float(by_method.get("yappy", 0))),
        ("Pago: sin método",      float(by_method.get("(sin pago)", 0))),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=v)
        if isinstance(v, float):
            c.number_format = MONEY_FMT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    # ---- Daily ----
    ws = wb.create_sheet("Daily")
    write_header(ws, ["Fecha", "Pedidos", "Cerrados", "Cancelados",
                      "Subtotal", "Propinas", "Total"])
    empty = lambda: {"orders": 0, "closed": 0, "cancelled": 0,
                     "subtotal": 0.0, "tip": 0.0, "total": 0.0}
    by_day = defaultdict(empty)
    for o in orders:
        dt = to_panama_dt(o.get("createdAt"))
        if not dt: continue
        b = by_day[dt.date()]
        b["orders"] += 1
        if o.get("status") == "closed":    b["closed"]    += 1
        if o.get("status") == "cancelled": b["cancelled"] += 1
        b["subtotal"] += o.get("subtotal",  0) or 0
        b["tip"]      += o.get("tipAmount", 0) or 0
        b["total"]    += o.get("total",     0) or 0

    cur, row = start_d, 2
    while cur <= end_d:
        b = by_day.get(cur, empty())
        ws.cell(row=row, column=1, value=cur.isoformat())
        ws.cell(row=row, column=2, value=b["orders"])
        ws.cell(row=row, column=3, value=b["closed"])
        ws.cell(row=row, column=4, value=b["cancelled"])
        ws.cell(row=row, column=5, value=b["subtotal"]).number_format = MONEY_FMT
        ws.cell(row=row, column=6, value=b["tip"]).number_format      = MONEY_FMT
        ws.cell(row=row, column=7, value=b["total"]).number_format    = MONEY_FMT
        cur += timedelta(days=1)
        row += 1
    autosize(ws)

    # ---- Orders ----
    ws = wb.create_sheet("Orders")
    write_header(ws, ["ID", "Fecha (Panamá)", "Mesa", "Estado", "# Items",
                      "Subtotal", "Tax", "Propina", "Total",
                      "Método pago", "Estado pago", "Pagado en", "Notas"])
    for r, o in enumerate(orders, 2):
        pay = o.get("payment") or {}
        dt      = to_panama_dt(o.get("createdAt"))
        paid_dt = to_panama_dt(pay.get("paidAt"))
        ws.cell(row=r, column=1,  value=o.get("_id"))
        ws.cell(row=r, column=2,  value=dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "")
        ws.cell(row=r, column=3,  value=o.get("tableNumber"))
        ws.cell(row=r, column=4,  value=o.get("status"))
        ws.cell(row=r, column=5,  value=o.get("itemCount"))
        ws.cell(row=r, column=6,  value=o.get("subtotal",  0) or 0).number_format = MONEY_FMT
        ws.cell(row=r, column=7,  value=o.get("taxAmount", 0) or 0).number_format = MONEY_FMT
        ws.cell(row=r, column=8,  value=o.get("tipAmount", 0) or 0).number_format = MONEY_FMT
        ws.cell(row=r, column=9,  value=o.get("total",     0) or 0).number_format = MONEY_FMT
        ws.cell(row=r, column=10, value=pay.get("method"))
        ws.cell(row=r, column=11, value=pay.get("status"))
        ws.cell(row=r, column=12, value=paid_dt.strftime("%Y-%m-%d %H:%M:%S") if paid_dt else "")
        ws.cell(row=r, column=13, value=o.get("notes"))
    autosize(ws)

    # ---- Items ----
    ws = wb.create_sheet("Items")
    write_header(ws, ["ID", "OrderID", "Fecha pedido", "Mesa", "Producto",
                      "Categoría", "Cantidad", "Precio unit", "Total",
                      "Modificadores", "Estado", "Instrucciones"])
    order_dt = {o["_id"]: to_panama_dt(o.get("createdAt")) for o in orders}
    for r, it in enumerate(items, 2):
        mods = it.get("modifiers") or []
        mods_str = "; ".join(f'{m.get("name","?")}={m.get("value","?")}' for m in mods)
        odt = order_dt.get(it.get("orderId"))
        ws.cell(row=r, column=1,  value=it.get("_id"))
        ws.cell(row=r, column=2,  value=it.get("orderId"))
        ws.cell(row=r, column=3,  value=odt.strftime("%Y-%m-%d %H:%M:%S") if odt else "")
        ws.cell(row=r, column=4,  value=it.get("tableNumber"))
        ws.cell(row=r, column=5,  value=it.get("productName"))
        ws.cell(row=r, column=6,  value=cat_map.get(it.get("categoryId"), ""))
        ws.cell(row=r, column=7,  value=it.get("quantity"))
        ws.cell(row=r, column=8,  value=it.get("unitPrice",  0) or 0).number_format = MONEY_FMT
        ws.cell(row=r, column=9,  value=it.get("totalPrice", 0) or 0).number_format = MONEY_FMT
        ws.cell(row=r, column=10, value=mods_str)
        ws.cell(row=r, column=11, value=it.get("status"))
        ws.cell(row=r, column=12, value=it.get("specialInstructions"))
    autosize(ws)

    # ---- TopProducts ----
    ws = wb.create_sheet("TopProducts")
    write_header(ws, ["Producto", "Cantidad", "Ingresos", "% del total"])
    by_prod = defaultdict(lambda: {"qty": 0, "revenue": 0.0})
    total_rev = 0.0
    for it in items:
        if it.get("status") == "cancelled":
            continue
        name = it.get("productName") or "?"
        by_prod[name]["qty"]     += it.get("quantity",  0) or 0
        by_prod[name]["revenue"] += it.get("totalPrice", 0) or 0
        total_rev                += it.get("totalPrice", 0) or 0
    ranked = sorted(by_prod.items(), key=lambda x: x[1]["revenue"], reverse=True)[:25]
    for r, (name, b) in enumerate(ranked, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=b["qty"])
        ws.cell(row=r, column=3, value=b["revenue"]).number_format = MONEY_FMT
        pct = (b["revenue"] / total_rev) if total_rev else 0
        ws.cell(row=r, column=4, value=pct).number_format = "0.0%"
    autosize(ws)

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)


# ---------- Main ----------

def main():
    p = argparse.ArgumentParser(description="Export Firestore order data to .xlsx")
    p.add_argument("--date",          help="Single day YYYY-MM-DD (Panamá TZ)")
    p.add_argument("--from", dest="from_d", help="Start date YYYY-MM-DD inclusive")
    p.add_argument("--to",   dest="to_d",   help="End date YYYY-MM-DD inclusive")
    p.add_argument("--out",           help="Output .xlsx path")
    args = p.parse_args()

    if args.date:
        start_d = end_d = parse_date(args.date)
    elif args.from_d or args.to_d:
        if not (args.from_d and args.to_d):
            sys.exit("Use both --from and --to")
        start_d = parse_date(args.from_d)
        end_d   = parse_date(args.to_d)
    else:
        today = datetime.now(PANAMA_TZ).date()
        start_d = today - timedelta(days=6)
        end_d   = today

    if start_d > end_d:
        sys.exit(f"--from ({start_d}) must be <= --to ({end_d})")

    out_path = args.out or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "reports",
        f"restaurant-os_{start_d}_{end_d}.xlsx"
    )

    print(f"Range : {start_d} → {end_d} (Panamá)")
    print(f"Output: {out_path}")

    token = get_token()
    if not token:
        sys.exit("ERROR: gcloud token missing. Run: gcloud auth login")

    print("Fetching orders…")
    orders = fetch_orders(start_d, end_d, token)
    print(f"  {len(orders)} orders")

    order_ids = [o["_id"] for o in orders]
    print("Fetching order_items…")
    items = fetch_items_for_orders(order_ids, token) if order_ids else []
    print(f"  {len(items)} items")

    print("Fetching categories…")
    categories = [parse_doc(d) for d in list_all("categories", token)]
    cat_map = {c["_id"]: c.get("name") for c in categories}
    print(f"  {len(cat_map)} categories")

    print("Building Excel…")
    build_workbook(orders, items, cat_map, start_d, end_d, out_path)
    print(f"Done: {out_path}")


if __name__ == "__main__":
    main()
